from openpyxl import load_workbook

workbook = load_workbook(filename='census/census.xlsx')

sheet_object = workbook.active
result_population = {}
current_county = sheet_object.cell(row=2, column=3).value
total_population = 0
for i in range(2, sheet_object.max_row):
    county = sheet_object.cell(row=i, column=3).value
    nb_population_county = sheet_object.cell(row=i, column=4).value
    if current_county == county:
        total_population += nb_population_county
    else:
        result_population[current_county] = total_population
        total_population = 0
        current_county = county

for item in result_population:
    print(f"{item}: {result_population[item]} habitants")
